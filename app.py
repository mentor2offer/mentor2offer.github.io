from flask import Flask, render_template, request, redirect, url_for, flash
from openpyxl import load_workbook, Workbook
import os

app = Flask(__name__)
app.secret_key = 'WsiRXWS7lNY0PVI'  # Replace with a real secret key

EXCEL_FILE = 'mentor2offer_mentorlist.xlsx'

def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(['Role', 'Email'])
        wb.save(EXCEL_FILE)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/join_waitlist', methods=['POST'])
def join_waitlist():
    role = request.form['role']
    email = request.form['email']
    
    if not email:
        flash('Please enter a valid email address.')
        return redirect(url_for('index'))
    
    init_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([role, email])
    wb.save(EXCEL_FILE)
    
    flash(f'Thank you! Your email ({email}) has been added to the {role} waitlist.')
    return redirect(url_for('index'))

if __name__ == '__main__':
    app.run(debug=True)